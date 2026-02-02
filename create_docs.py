from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta

# ワークブック作成
wb = Workbook()

# ===== シート1: 技術仕様書 =====
ws1 = wb.active
ws1.title = "技術仕様書"

# ヘッダースタイル
header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# タイトル
ws1['A1'] = "ガードプール 技術仕様書"
ws1['A1'].font = Font(size=16, bold=True, color="1E3A8A")
ws1.merge_cells('A1:D1')
ws1['A1'].alignment = center_align

ws1['A2'] = f"作成日: {datetime.now().strftime('%Y年%m月%d日')}"
ws1.merge_cells('A2:D2')
ws1['A2'].alignment = Alignment(horizontal="right")

# ヘッダー行
ws1['A4'] = "項目"
ws1['B4'] = "内容"
ws1['C4'] = "詳細"
ws1['D4'] = "URL/備考"

for col in ['A4', 'B4', 'C4', 'D4']:
    ws1[col].fill = header_fill
    ws1[col].font = header_font
    ws1[col].alignment = center_align
    ws1[col].border = border

# データ
data = [
    ["プロジェクト名", "ガードプール (GuardPool)", "警備員シェアリングシステム", ""],
    ["", "", "", ""],
    ["【フロントエンド】", "", "", ""],
    ["ホスティング", "Netlify", "自動デプロイ対応", "https://ats-guardpool.netlify.app"],
    ["公開URL", "https://ats-guardpool.netlify.app", "本番環境", ""],
    ["HTML", "index.html", "ログイン画面", ""],
    ["", "dashboard.html", "ダッシュボード", ""],
    ["CSS", "style.css", "デザイン・レイアウト", "ネイビーブルー + ピンク配色"],
    ["JavaScript", "app.js", "ログイン・画面制御", "Vanilla JavaScript"],
    ["", "", "", ""],
    ["【バックエンド】", "", "", ""],
    ["データベース", "Supabase", "PostgreSQL", "https://supabase.com"],
    ["プロジェクト名", "guardpool", "本番環境", ""],
    ["リージョン", "Northeast Asia (Tokyo)", "東京リージョン", ""],
    ["", "", "", ""],
    ["【データベーステーブル】", "", "", ""],
    ["companies", "会社マスター", "10社登録済み", "ログインID/パスワード管理"],
    ["posts_available", "人が余ってます投稿", "未実装", ""],
    ["posts_needed", "人が足りません投稿", "未実装", ""],
    ["matches", "マッチング履歴", "未実装", ""],
    ["", "", "", ""],
    ["【ソースコード管理】", "", "", ""],
    ["Git リポジトリ", "GitHub", "バージョン管理", "https://github.com/hondaporta-ship-it/guardpool"],
    ["ブランチ", "main", "本番ブランチ", ""],
    ["", "", "", ""],
    ["【開発環境】", "", "", ""],
    ["エディタ", "Visual Studio Code", "コード編集", ""],
    ["ターミナル", "macOS Terminal", "コマンド実行", ""],
    ["Python", "Python 3.x", "スクリプト実行", "openpyxl使用"],
    ["", "", "", ""],
    ["【認証】", "", "", ""],
    ["ログイン方式", "ID/パスワード認証", "シンプル認証", ""],
    ["セッション管理", "sessionStorage", "ブラウザ内保存", ""],
    ["", "", "", ""],
    ["【初期データ】", "", "", ""],
    ["登録済み会社", "10社", "福岡エリア警備会社", ""],
    ["ATS", "login_id: ats", "password: ats2025", "テスト用アカウント"],
    ["その他9社", "各社ID/パスワード設定済み", "alpha, zenkyushu, thanks等", ""],
    ["", "", "", ""],
    ["【セキュリティ】", "", "", ""],
    ["HTTPS", "対応", "Netlify標準", ""],
    ["パスワード", "平文保存（暫定）", "将来的にハッシュ化推奨", ""],
    ["アクセス制限", "ログイン必須", "未ログインは index.html へリダイレクト", ""],
]

row = 5
for item in data:
    ws1[f'A{row}'] = item[0]
    ws1[f'B{row}'] = item[1]
    ws1[f'C{row}'] = item[2]
    ws1[f'D{row}'] = item[3]
    
    for col in ['A', 'B', 'C', 'D']:
        cell = ws1[f'{col}{row}']
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        if item[0].startswith("【"):
            cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
            cell.font = Font(bold=True, color="1E3A8A")
    
    row += 1

# 列幅調整
ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 25
ws1.column_dimensions['C'].width = 30
ws1.column_dimensions['D'].width = 40

# ===== シート2: 開発スケジュール =====
ws2 = wb.create_sheet("開発スケジュール")

# タイトル
ws2['A1'] = "ガードプール 開発スケジュール"
ws2['A1'].font = Font(size=16, bold=True, color="1E3A8A")
ws2.merge_cells('A1:E1')
ws2['A1'].alignment = center_align

ws2['A2'] = "1週間でローンチまでの流れ"
ws2['A2'].font = Font(size=12, bold=True, color="EF4444")
ws2.merge_cells('A2:E2')
ws2['A2'].alignment = center_align

# 現在の状況
ws2['A4'] = "【現在の状況】"
ws2['A4'].font = Font(bold=True, size=11)
ws2['A4'].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
ws2['A4'].font = Font(color="FFFFFF", bold=True)
ws2.merge_cells('A4:E4')

status_data = [
    ["✅ 完了", "ログイン機能", "動作確認済み", "", ""],
    ["✅ 完了", "ダッシュボード表示", "会社名表示", "", ""],
    ["✅ 完了", "データベース構築", "Supabase 4テーブル作成", "", ""],
    ["✅ 完了", "初期データ投入", "10社登録完了", "", ""],
    ["✅ 完了", "デプロイ", "Netlify 公開済み", "", ""],
    ["❌ 未完了", "投稿機能", "開発中", "", ""],
    ["❌ 未完了", "投稿一覧表示", "開発中", "", ""],
]

row = 5
for item in status_data:
    for col_idx, value in enumerate(item):
        col = chr(65 + col_idx)
        ws2[f'{col}{row}'] = value
        ws2[f'{col}{row}'].border = border
        if item[0] == "✅ 完了":
            ws2[f'A{row}'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        else:
            ws2[f'A{row}'].fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    row += 1

# スケジュール
ws2[f'A{row+1}'] = "【今後のスケジュール】"
ws2[f'A{row+1}'].font = Font(bold=True, size=11)
ws2[f'A{row+1}'].fill = header_fill
ws2[f'A{row+1}'].font = Font(color="FFFFFF", bold=True)
ws2.merge_cells(f'A{row+1}:E{row+1}')

row += 2

# ヘッダー
headers = ["日付", "タスク", "担当", "所要時間", "備考"]
for col_idx, header in enumerate(headers):
    col = chr(65 + col_idx)
    ws2[f'{col}{row}'] = header
    ws2[f'{col}{row}'].fill = header_fill
    ws2[f'{col}{row}'].font = header_font
    ws2[f'{col}{row}'].alignment = center_align
    ws2[f'{col}{row}'].border = border

row += 1

# 今日から7日間のスケジュール
base_date = datetime.now()
schedule_data = [
    [0, "水曜プレゼン（モック版）", "ヒデさん", "30分", "三津谷社長にデモ"],
    [0, "社長からのフィードバック収集", "ヒデさん", "15分", "改善点を確認"],
    [1, "投稿フォーム開発", "開発", "3時間", "余り・足りない両方"],
    [2, "投稿一覧表示機能", "開発", "2時間", "リアルタイム更新"],
    [3, "マッチング機能", "開発", "2時間", "興味ありボタン"],
    [4, "通知機能（簡易版）", "開発", "1時間", "システム内通知"],
    [4, "デザイン調整", "開発", "1時間", "UI/UX改善"],
    [5, "テスト運用開始", "ヒデさん", "1日", "ATS + 2社でテスト"],
    [6, "バグ修正・調整", "開発", "2時間", "フィードバック反映"],
    [7, "本番ローンチ", "ヒデさん", "1日", "10社に案内・稼働開始"],
]

for item in schedule_data:
    date = base_date + timedelta(days=item[0])
    date_str = date.strftime("%m/%d (%a)")
    
    ws2[f'A{row}'] = date_str
    ws2[f'B{row}'] = item[1]
    ws2[f'C{row}'] = item[2]
    ws2[f'D{row}'] = item[3]
    ws2[f'E{row}'] = item[4]
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws2[f'{col}{row}'].border = border
        ws2[f'{col}{row}'].alignment = Alignment(vertical="top", wrap_text=True)
    
    # 今日の行をハイライト
    if item[0] == 0:
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws2[f'{col}{row}'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    
    row += 1

# まとめ
ws2[f'A{row+1}'] = "【まとめ】"
ws2[f'A{row+1}'].font = Font(bold=True, size=11)
ws2[f'A{row+1}'].fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
ws2[f'A{row+1}'].font = Font(color="FFFFFF", bold=True)
ws2.merge_cells(f'A{row+1}:E{row+1}')

row += 2

summary = [
    ["✓", "水曜日", "モック版でプレゼン → 社長の反応を確認", "", ""],
    ["✓", "木〜土曜", "投稿機能・一覧表示を開発", "", ""],
    ["✓", "日曜", "テスト運用（3社）", "", ""],
    ["✓", "翌週月曜", "本番ローンチ（10社）", "", ""],
    ["", "", "", "", ""],
    ["目標", "1週間後", "福岡10社で稼働開始", "", "月額30万円の収入見込み"],
]

for item in summary:
    for col_idx, value in enumerate(item):
        col = chr(65 + col_idx)
        ws2[f'{col}{row}'] = value
        ws2[f'{col}{row}'].border = border
        ws2[f'{col}{row}'].alignment = Alignment(vertical="top", wrap_text=True)
        if item[0] == "目標":
            ws2[f'{col}{row}'].fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            ws2[f'A{row}'].font = Font(bold=True)
    row += 1

# 列幅調整
ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 35

# ファイル保存
filename = "ガードプール_技術仕様書_スケジュール.xlsx"
wb.save(filename)
print(f"✅ ファイル作成完了: {filename}")
print(f"📁 保存場所: /Users/hidekihonda/guardpool/{filename}")